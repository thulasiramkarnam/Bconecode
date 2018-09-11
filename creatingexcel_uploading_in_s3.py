

import xlrd
import xlwt
import boto3
import logging
import json
import openpyxl 
logger = logging.getLogger()
logger.setLevel(logging.INFO)

def lambda_handler(event, context):
    # TODO implement
    ab = {'glass': ['silica', 'zinc', 'mica'], 'tyre': ['mclean','aws']}
    wb = openpyxl.Workbook() 
    counter = 0
    for k,v in ab.items():
        
        sheet = wb.create_sheet(index = counter , title = k)
        for i in range(len(v)):
            sheet.cell(row = i+1, column = 1).value = v[i]
        counter += counter
    wb.save("/tmp/demo.xlsx") 
    # workbook = xlwt.Workbook(encoding = 'ascii')
    # worksheet = workbook.add_sheet('My Worksheet')
    # worksheet.write(0, 0, label = 'Row 0, Column 0 34')
    # logger.info("before saving")
    # workbook.save('Excel_Workbook.xls')
    # logger.info("After saving")
    # with open('/tmp/Excel_Workbook.xls', 'wb') as file:
    s3 = boto3.resource('s3')
    s3.meta.client.upload_file('/tmp/demo.xlsx', 'neo-apps-procoure.ai', 'xlrdxlwt.xlsx')
    # logger.info(event)
    # logger.info(type(event))
    # logger.info("After event")
    # return {
    #     "statusCode": 200,
    #     "body": json.dumps('Hello from Lambda!')
    # }
