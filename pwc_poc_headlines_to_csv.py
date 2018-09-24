
import requests
import logging
import boto3
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Color, colors
import json
import datetime
from datetime import datetime as dtt
logger = logging.getLogger()
logger.setLevel(logging.INFO)
def lambda_handler(event, context):
    url = ('https://newsapi.org/v2/top-headlines?'
        'sources=bbc-news,al-jazeera-english,cnn&'
        'apiKey=559f97a41d574b989609df347b5adfa8')
    cd = requests.get(url)
    logger.info(cd.text)
    logger.info("After output from lambda")
    headlines = json.loads(cd.text)['articles']
    file_name_in_s3 = "Headlines_at_"+(str(dtt.now()))+".csv"
    file_name = "Headlines_at_"+(str(dtt.now()))+".csv"
    file_path = '/tmp/' + file_name
    
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    alignment=Alignment(wrap_text=True)
    with open(file_path, 'wb') as file_name:
        book = Workbook()
        sheet = book.active
        ft = Font(color=colors.BLACK,bold=True)
        sheet.title = "Headlines"
        sheet['A1'] = "Source"
        sheet['A1'].font = ft
        sheet.column_dimensions['A'].width = 28
        sheet['B1'] = "Timestamp"
        sheet['B1'].font = ft
        sheet.column_dimensions['B'].width = 28
        sheet['C1'] = "PublishedAt"
        sheet['C1'].font = ft
        sheet.column_dimensions['C'].width = 28
        sheet['D1'] = "Headline"
        sheet['D1'].font = ft
        sheet.column_dimensions['D'].width = 50
        counter = 2
        for length,value in enumerate(headlines, start= 2):
            headline = []
            headline.extend([value['source']['name'], str(dtt.now()), value['publishedAt'],value['description']])
            for i,j in enumerate(headline, start = 1):
                sheet.cell(row=length, column=i).value = j
                sheet.cell(row=length, column=i).border = thin_border
                if headline[-1] == j:
                    val = 'D'+str(counter)
                    print (val)
                    print ("after val")
                    sheet[val].alignment = alignment
            counter += 1
        book.save(file_name)
    s3 = boto3.resource('s3')
    s3.meta.client.upload_file(file_path, 'neo-apps-procoure.ai', file_name_in_s3)
    logger.info("uploaded successfully")
